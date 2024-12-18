import tkinter as tk
from tkinter import messagebox, ttk
import customtkinter as ctk
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import pandas as pd
from tkinter import filedialog
import openpyxl
import numpy as np
from datetime import datetime
from tkcalendar import DateEntry
import sqlite3
style = ttk.Style()
VILLES_MAROC = {
    'Casablanca': (33.5731, -7.5898),
    'Rabat': (34.0209, -6.8416), 
    'Marrakech': (31.6295, -7.9811),
    'Fès': (34.0333, -5.0000),
    'Tanger': (35.7595, -5.8340),
    'Agadir': (30.4278, -9.5981),
    'Meknès': (33.8731, -5.5407),
    'Oujda': (34.6867, -1.9114),
    'Kenitra': (34.2610, -6.5802),
    'Tetouan': (35.5784, -5.3630),
    'Safi': (32.2994, -9.2372),
    'El Jadida': (33.2316, -8.5007),
    'Nador': (35.1667, -2.9333),
    'Beni Mellal': (32.3373, -6.3498),
    'Mohammedia': (33.6833, -7.3833),
    'Taza': (34.2167, -4.0167),
    'Khouribga': (32.8833, -6.9167),
    'Settat': (33.0000, -7.6167),
    'Errachidia': (31.9333, -4.4167),
    'Larache': (35.1833, -6.1500),
    'Khemisset': (33.8167, -6.0667),
    'Ouarzazate': (30.9167, -6.8833),
    'Tiznit': (29.7000, -9.7333),
    'Tan-Tan': (28.4333, -11.1000),
    'Guelmim': (28.9833, -10.0667),
    'Ifrane': (33.5333, -5.1000),
    'Asilah': (35.4667, -6.0333),
    'Chefchaouen': (35.1667, -5.2667),
    'Al Hoceima': (35.2500, -3.9333),
    'Taroudant': (30.4667, -8.8833),
    'Oued Zem': (32.8667, -6.5667),
    'Azrou': (33.4333, -5.2167),
    'Sidi Kacem': (34.2167, -5.7000),
    'Sidi Slimane': (34.2667, -5.9167),
    'Boujdour': (26.1333, -14.4833),
    'Laâyoune': (27.1500, -13.2000),
    'Dakhla': (23.7167, -15.9333)
}
style.configure(
    "TCombobox",
    background="#f0f0f0",
    fieldbackground="white",
    foreground="#333",
    arrowcolor="#333"
)

plt.style.use('ggplot')  # ou 'fivethirtyeight', 'bmh', 'classic'
plt.rcParams.update({
    'font.family': 'sans-serif',
    'font.sans-serif': ['Arial'],
    'font.size': 10,
    'axes.titlesize': 12,
    'axes.labelsize': 11,
    'figure.facecolor': 'white',
    'axes.facecolor': 'white',
    'axes.grid': True,
    'grid.alpha': 0.3
})

# Style pour ttk
style = ttk.Style()
style.configure(
    "TCombobox",
    background="#f0f0f0",
    fieldbackground="white",
    foreground="#333",
    arrowcolor="#333"
)

# Style pour Treeview
style.configure(
    "Treeview",
    background="white",
    fieldbackground="white",
    foreground="black",
    rowheight=30
)
style.configure(
    "Treeview.Heading",
    background="#f0f0f0",
    foreground="#333",
    font=('Arial', 10, 'bold')
)
style.map(
    "Treeview",
    background=[('selected', '#2196F3')],
    foreground=[('selected', 'white')]
)
class AdminPage(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master)
        self.after_ids = []  # Ajoutez cette ligne
        self.configure(fg_color="white")
        

        # Couleurs
        self.orange = "#ff8c00"
        self.blue_ciel = "#87cefa"
        self.bg_color = "#f2f2f2"
        self.blue_ciel = "#2196F3"
        self.orange = "#FF9800"
        self.green = "#4CAF50"
        self.red = "#F44336"
        

        # Barre supérieure
        self.framesup = ctk.CTkFrame(self, height=70, corner_radius=0, fg_color=self.blue_ciel)
        self.framesup.pack(fill="x", side="top")
        self.speedy_label = ctk.CTkLabel(self.framesup, text="Administration", font=("Helvetica", 50, "bold"), text_color="#ff8c00")
        self.speedy_label.pack(side="left",padx=5,pady=(10, 30))

        # Agencement principal
        main_frame = ctk.CTkFrame(self, fg_color="#f2f2f2", bg_color="white")
        main_frame.pack(fill="both", expand=True)

        # Barre latérale
        sidebar_frame = ctk.CTkScrollableFrame(main_frame, fg_color=self.blue_ciel,width=300)
        sidebar_frame.pack(side="left", fill="y", padx=0, pady=10)

        # Contenu principal
        self.content_frame = ctk.CTkScrollableFrame(main_frame,fg_color="white")
        self.content_frame.pack(side="right", fill="both", expand=True, padx=10)

        # Boutons dans la barre latérale
        self.create_sidebar_button(sidebar_frame, "👥 Gestion des clients", self.manage_clients)
        self.create_sidebar_button(sidebar_frame, "📦 Suivi des colis", self.tracking_packages)
        self.create_sidebar_button(sidebar_frame, "📊 Statistiques", self.view_statistics)
        self.create_sidebar_button(sidebar_frame, "🔔 Envoi des notifications", self.send_notifications)
        self.create_sidebar_button(sidebar_frame, "👨‍💼 Gestion des employés", self.employee)
        self.create_sidebar_button(sidebar_frame, "🚚  Gestion des véhicules", self.manage_vehicles)
        self.create_sidebar_button(sidebar_frame, "🏠 Demandes Déménagement", self.manage_moving_requests)
        self.create_sidebar_button(sidebar_frame, "🚪 Déconnexion", self.logout)
        self.show_welcome()
        self.db_path = "suivi_coli.db"
        try:
            conn = sqlite3.connect(self.db_path)
            print("Connexion à la base de données établie avec succès")
            
        except sqlite3.Error as e:
            print(f"Erreur de connexion à la base de données: {e}")
            messagebox.showerror("Erreur", f"Erreur de connexion à la base de données: {e}")    
            
    def show_welcome(self):
        # Nettoyer le contenu actuel
        self.clear_content()
        
        # Frame principal
        welcome_frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        welcome_frame.pack(fill="both", expand=True, padx=40, pady=40)
        
        # Message de bienvenue
        welcome_label = ctk.CTkLabel(
            welcome_frame,
            text="👋 Bienvenue dans votre Espace Administration",
            font=("Arial", 32, "bold"),
            text_color=self.orange
        )
        welcome_label.pack(pady=(0, 30))

        # Sous-titre
        subtitle = ctk.CTkLabel(
            welcome_frame,
            text="Guide d'utilisation des fonctionnalités",
            font=("Arial", 24, "bold"),
            text_color=self.blue_ciel
        )
        subtitle.pack(pady=(0, 20))

        # Guide des fonctionnalités
        features = [
            ("👥 Gestion des clients", "Gérez votre base de clients, ajoutez, modifiez ou supprimez des informations client"),
            ("📦 Suivi des colis", "Suivez en temps réel l'état et la position des colis en cours de livraison"),
            ("📊 Statistiques", "Visualisez les statistiques et analyses de performance"),
            ("🔔 Envoi des notifications", "Envoyez des notifications aux clients et aux livreurs"),
            ("👨‍💼 Gestion des employés", "Gérez les informations et les accès des employés"),
            ("🚚 Gestion des véhicules", "Suivez et gérez votre flotte de véhicules"),
            ("🏠 Demandes Déménagement", "Gérez les demandes de déménagement des clients"),
            ("🚪 Déconnexion", "Déconnectez-vous de votre session")
        ]

        # Créer un frame scrollable pour les fonctionnalités
        features_frame = ctk.CTkScrollableFrame(
            welcome_frame,
            fg_color="transparent",
            height=400
        )
        features_frame.pack(fill="x", pady=20)

        # Ajouter chaque fonctionnalité avec son explication
        for icon_title, description in features:
            feature_frame = ctk.CTkFrame(features_frame, fg_color=self.bg_color)
            feature_frame.pack(fill="x", pady=5, padx=10)
            
            title = ctk.CTkLabel(
                feature_frame,
                text=icon_title,
                font=("Arial", 16, "bold"),
                text_color=self.orange
            )
            title.pack(anchor="w", padx=10, pady=(5, 0))
            
            desc = ctk.CTkLabel(
                feature_frame,
                text=description,
                font=("Arial", 12),
                text_color="gray",
                wraplength=800
            )
            desc.pack(anchor="w", padx=10, pady=(0, 5))

        # Note de bas de page
        footer_note = ctk.CTkLabel(
            welcome_frame,
            text="Pour commencer, cliquez sur l'une des options du menu à gauche",
            font=("Arial", 14, "italic"),
            text_color="gray"
        )
        footer_note.pack(pady=20)
    def manage_moving_requests(self):
        self.clear_content()
        
        # Frame principal avec titre
        title_frame = ctk.CTkFrame(self.content_frame, fg_color=self.blue_ciel, height=60)
        title_frame.pack(fill="x", pady=(0, 20))
        
        # Titre
        title_label = ctk.CTkLabel(
            title_frame,
            text="Gestion des Déménagements",
            font=("Arial", 24, "bold"),
            text_color="white"
        )
        title_label.pack(pady=10)

        # Frame pour la barre de recherche
        search_frame = ctk.CTkFrame(self.content_frame)
        search_frame.pack(fill="x", padx=10, pady=5)

        # Barre de recherche par ID
        search_label = ctk.CTkLabel(search_frame, text="Rechercher par ID:", font=("Arial", 12))
        search_label.pack(side="left", padx=5)
        
        
        search_entry = ctk.CTkEntry(search_frame, width=100)
        search_entry.pack(side="left", padx=5)
        def search_by_id():
          """Rechercher une demande par ID"""
          search_id=search_entry.get()
          
          if not search_id:
            messagebox.showwarning("Attention", "Veuillez entrer un ID")
            return
            
          found = False
          for item in self.tree.get_children():
            values = self.tree.item(item)['values']
            if str(values[0]) == search_id:
                # Sélectionner et mettre en évidence l'élément trouvé
                self.tree.selection_set(item)
                self.tree.see(item)
                found = True
                break
        
          if not found:
            messagebox.showwarning("Recherche", "Aucune demande trouvée avec cet ID")
        
        search_button = ctk.CTkButton(
            search_frame,
            text="Rechercher",
            command= search_by_id,
            width=100
        )
        search_button.pack(side="left", padx=5)

        # Frame pour le tableau
        table_frame = ctk.CTkFrame(self.content_frame)
        table_frame.pack(fill="both", expand=True, padx=10, pady=5)

        # Tableau des déménagements
        columns = ("ID", "Adresse départ", "Adresse arrivée", "Date de demenage", "Heure de demenage", "objets", "nombre de cartons", "nombre de personnes", "nom du service", "État")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=15)
        
        # Configuration des colonnes avec des largeurs réduites
        column_widths = {
            "ID": 50,
            "Adresse départ": 120,
            "Adresse arrivée": 120,
            "Date de demenage": 100,
            "Heure de demenage": 100,
            "objets": 100,
            "nombre de cartons": 80,
            "nombre de personnes": 80,
            "nom du service": 100,
            "État": 80
        }
        
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=column_widths[col], anchor="center")

        # Scrollbar
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        # Placement du tableau et scrollbar
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Frame séparé pour les boutons APRÈS le tableau
        button_frame = ctk.CTkFrame(self.content_frame)
        button_frame.pack(fill="x", padx=10, pady=10)

        # Boutons
        accept_button = ctk.CTkButton(
            button_frame,
            text="Accepter",
            command=self.change_status_to_accepted,
            fg_color="green",
            width=120,
            height=32
        )
        accept_button.pack(side="left", padx=5)

        reject_button = ctk.CTkButton(
            button_frame,
            text="Refuser",
            command=self.change_status_to_rejected,
            fg_color="red",
            width=120,
            height=32
        )
        reject_button.pack(side="left", padx=5)

        # Charger les données
        self.load_sample_datadem()
    
    def change_status_to_accepted(self):
        """Changer le statut à 'Accepté'"""
        selected_items = self.tree.selection()
        if not selected_items:
            
            return
            
        try:
            conn = sqlite3.connect("suivi_coli.db")
            cursor = conn.cursor()
            
            for item in selected_items:
                values = list(self.tree.item(item)['values'])
                if values[9] != "Accepté":
                    # Mettre à jour la base de données
                    cursor.execute("""
                        UPDATE demenagement 
                        SET etat = 'Accepté'
                        WHERE id_dem = ?
                    """, (values[0],))
                    
                    # Mettre à jour l'interface
                    values[9] = "Accepté"
                    self.tree.item(item, values=values)
                    
            
            conn.commit()
            
        except sqlite3.Error as e:
            pass
        finally:
            if 'conn' in locals():
                conn.close()

    def change_status_to_rejected(self):
        """Changer le statut à 'Refusé'"""
        selected_items = self.tree.selection()
        if not selected_items:
            
            return
            
        try:
            conn = sqlite3.connect("suivi_coli.db")
            cursor = conn.cursor()
            
            for item in selected_items:
                values = list(self.tree.item(item)['values'])
                if values[9] != "Refusé":
                    # Mettre à jour la base de données
                    cursor.execute("""
                        UPDATE demenagement 
                        SET etat = 'Refusé'
                        WHERE id_dem = ?
                    """, (values[0],))
                    
                    # Mettre à jour l'interface
                    values[9] = "Refusé"
                    self.tree.item(item, values=values)
                    
            
            conn.commit()
            
        except sqlite3.Error as e:
            pass 
        finally:
            if 'conn' in locals():
                conn.close()
    def load_sample_datatr(self):
    # Nettoyer les données existantes
     for item in self.tracking_tree.get_children():
        self.tracking_tree.delete(item)
            
     try:
        # Connexion à la base de données
        conn = sqlite3.connect("suivi_coli.db")
        cursor = conn.cursor()

        # Requête SQL pour récupérer les informations nécessaires
        query = """
        SELECT 
            c.id_colis,
            u.id_user,
            c.posit_depart_lat,
            c.posit_depart_lon,
            CASE c.etatlivrs
                WHEN 0 THEN 'En attente'
                WHEN 1 THEN 'En cours'
                WHEN 2 THEN 'Livré'
                ELSE 'Inconnu'
            END as etat,
            c.posit_depart_lat,
            c.posit_depart_lon,
            c.date_livrs
        FROM colis c
        LEFT JOIN user u ON c.id_user = u.id_user
        """

        cursor.execute(query)
        colis_data = cursor.fetchall()

        # Insérer les données dans le tableau
        for colis in colis_data:
            self.tracking_tree.insert('', 'end', values=colis)

        # Si aucune donnée trouvée
        if not colis_data:
            self.tracking_tree.insert('', 'end', values=('Aucun colis trouvé', '', '', '', '', '', '', ''))

     except sqlite3.Error as e:
        messagebox.showerror("Erreur", f"Erreur lors du chargement des colis : {e}")
        self.tracking_tree.insert('', 'end', values=('Erreur de chargement', '', '', '', '', '', '', ''))
    
     finally:
        if 'conn' in locals():
            conn.close()
        

    def filter_requests(self):
        """Filtrer les demandes selon la recherche et le statut"""
        search_term = self.search_var.get().lower()
        status = self.status_var.get()
        
        for item in self.tree.get_children():
            values = self.tree.item(item)['values']
            show = True
            
            if search_term:
                show = any(search_term in str(value).lower() for value in values)
            
            if status != "Tous":
                show = show and values[7] == status
            
            if show:
                self.tree.reattach(item, '', 'end')
            else:
                self.tree.detach(item)

    def reset_filters(self):
        """Réinitialiser les filtres"""
        self.search_var.set("")
        self.status_var.set("Tous")
        self.load_sample_datadem()  # Recharger les données

    def sort_column(self, col):
        """Trier le tableau par colonne"""
        l = [(self.tree.set(k, col), k) for k in self.tree.get_children('')]
        l.sort(reverse=getattr(self, '_sort_reverse', False))
        self._sort_reverse = not getattr(self, '_sort_reverse', False)
        
        for index, (_, k) in enumerate(l):
            self.tree.move(k, '', index)
    def load_sample_datadem(self):
        # Nettoyer les données existantes
        for item in self.tree.get_children():
            self.tree.delete(item)
                
        try:
            # Connexion à la base de données
            conn = sqlite3.connect("suivi_coli.db")
            cursor = conn.cursor()

            # Requête SQL pour récupérer les données de déménagement
            query = """
            SELECT 
                id_dem,
                adress_depart,
                adress_arrive,
                date_dem,
                heure_dem,
                objet_volum,
                nbr_cart,
                nbr_pers,
                nom_service,
                etat
            FROM demenagement
            """

            cursor.execute(query)
            dem_data = cursor.fetchall()

            # Insérer les données dans le tableau
            for dem in dem_data:
                self.tree.insert('', 'end', values=dem)

            # Si aucune donnée trouvée
            if not dem_data:
                self.tree.insert('', 'end', values=('Aucune demande trouvée', '', '', '', '', '', '', '', '', ''))

        except sqlite3.Error as e:
            messagebox.showerror("Erreur", f"Erreur lors du chargement des demandes de déménagement : {e}")
            self.tree.insert('', 'end', values=('Erreur de chargement', '', '', '', '', '', '', '', '', ''))
        
        finally:
            if 'conn' in locals():
                conn.close()



    def create_sidebar_button(self, frame, text, command):
        button = ctk.CTkButton(frame, text=text, font=("arial", 20, "bold"), command=command, width=55, height=60, corner_radius=8, fg_color=self.orange)
        button.pack(fill="x", pady=25)

    
    def manage_vehicles(self):
        self.clear_content()
    
    # Initialiser les variables avec self comme master
        self.vehicle_search_var = tk.StringVar(master=self)
        self.vehicle_state_var = tk.StringVar(master=self, value="Tous")
        
        # Frame titre
        title_frame = ctk.CTkFrame(self.content_frame, fg_color=self.blue_ciel, height=60)
        title_frame.pack(fill="x", pady=(0, 20))
        title_label = ctk.CTkLabel(
            title_frame,
            text="Gestion des Véhicules",
            font=("Arial", 24, "bold"),
            text_color="white"
        )
        title_label.pack(pady=10)

        # Frame recherche
        search_frame = ctk.CTkFrame(self.content_frame)
        search_frame.pack(fill="x", padx=10, pady=5)

        # Barre de recherche
        search_label = ctk.CTkLabel(search_frame, text="Rechercher:", font=("Arial", 12))
        search_label.pack(side="left", padx=5)
        
        self.vehicle_search_var = tk.StringVar()
        search_entry = ctk.CTkEntry(
            search_frame,
            textvariable=self.vehicle_search_var,
            width=200,
            placeholder_text="Matricule..."
        )
        search_entry.pack(side="left", padx=5)

        # Filtre par état
        state_label = ctk.CTkLabel(search_frame, text="État:", font=("Arial", 12))
        state_label.pack(side="left", padx=5)
        
        self.vehicle_state_var = tk.StringVar(value="Tous")
        state_combo = ttk.Combobox(
            search_frame,
            textvariable=self.vehicle_state_var,
            values=["Tous", "Disponible", "En mission", "En maintenance"],
            width=15,
            state="readonly"
        )
        state_combo.pack(side="left", padx=5)
        def search_vehicle():
         try:
          search_term = search_entry.get().lower()
          state_filter = state_combo.get()
          conn=sqlite3.connect("suivi_coli.db")
          cursor = conn.cursor()
          if state_filter == "Tous":
            # Si "Tous" est sélectionné, récupérer tous les véhicules
            query = """
                SELECT id_vehicul, marque, modele, année, état, kilometrage, 
                       dernieremaintenance 
                FROM vehicule
            """
          else:
            # Sinon, filtrer par état
            query = """
                SELECT id_vehicul, marque, modele, année, état, kilometrage, 
                       dernieremaintenance 
                FROM vehicule 
                WHERE état = ?
            """
            cursor.execute(query, (state_filter,))
          vehicles = cursor.fetchall()
          
          for item in self.vehicle_tree.get_children():
            self.vehicle_tree.delete(item)
          # Afficher les résultats
          for vehicle in vehicles:
            self.vehicle_tree.insert('', 'end', values=vehicle)  
          if not vehicles:
            messagebox.showinfo("Recherche", "Aucun véhicule trouvé")
         except sqlite3.Error as e:
           print(f"Erreur lors de la recherche: {e}")
           messagebox.showerror("Erreur", f"Erreur lors de la recherche: {e}")
         finally:
            if 'conn' in locals():
             conn.close()

        def reset_vehicle_search():
          search_entry.insert(0,"")
          state_combo.set("Tous")
          self.hicle_data()

        # Boutons de recherche et réinitialisation
        search_button = ctk.CTkButton(
            search_frame,
            text="🔍 Rechercher",
            command=search_vehicle,
            width=100,
            fg_color=self.orange
        )
        search_button.pack(side="left", padx=5)

        reset_button = ctk.CTkButton(
            search_frame,
            text="↺ Réinitialiser",
            command=reset_vehicle_search,
            width=100,
            fg_color=self.orange
        )
        reset_button.pack(side="left", padx=5)

        # Frame pour le tableau
        table_frame = ctk.CTkFrame(self.content_frame)
        table_frame.pack(fill="both", expand=True, padx=10, pady=5)

        # Tableau des véhicules
        columns = ("id_vehicul", "Marque", "Modèle", "Année", "État", "Kilométrage", "Dernière maintenance")
        self.vehicle_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=15)
        
        # Configuration des colonnes
        for col in columns:
            self.vehicle_tree.heading(col, text=col)
            width = 150 if col in ["Marque", "Modèle"] else 100
            self.vehicle_tree.column(col, width=width, anchor="center")

        # Scrollbar
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.vehicle_tree.yview)
        self.vehicle_tree.configure(yscrollcommand=scrollbar.set)
        
        # Placement du tableau et scrollbar
        self.vehicle_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Frame séparé pour les boutons APRÈS le tableau
        button_frame = ctk.CTkFrame(self.content_frame)
        button_frame.pack(fill="x", padx=10, pady=10)

        # Boutons d'action
        add_button = ctk.CTkButton(
            button_frame,
            text="➕ Ajouter",
            command=self.show_add_vehicle_form,
            fg_color="green",
            width=120,
            height=32
        )
        add_button.pack(side="left", padx=5)

        modify_button = ctk.CTkButton(
            button_frame,
            text="✏️ Modifier",
            command=self.show_modify_vehicle_form,
            fg_color="orange",
            width=120,
            height=32
        )
        modify_button.pack(side="left", padx=5)

        delete_button = ctk.CTkButton(
            button_frame,
            text="🗑️ Supprimer",
            command=self.delete_vehicle,
            fg_color="red",
            width=120,
            height=32
        )
        delete_button.pack(side="left", padx=5)
        export_button = ctk.CTkButton(
        button_frame,
        text="📊 Exporter Excel",
        command=self.export_to_excelveh,
        fg_color="#2E7D32",  # Vert foncé
        width=120,
        height=32
    )
        export_button.pack(side="left", padx=5)


    # Charger les données
        self.load_vehicle_data()


    def export_to_excelveh(self):
        try:
            # Récupérer toutes les données du tableau
            data = []
            columns = ["Matricule", "Marque", "Modèle", "Année", "État", "Kilométrage", "Dernière maintenance"]
            
            for item in self.vehicle_tree.get_children():
                values = self.vehicle_tree.item(item)['values']
                data.append(values)

            # Créer un DataFrame pandas
            df = pd.DataFrame(data, columns=columns)

            # Demander à l'utilisateur où sauvegarder le fichier
            file_path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Enregistrer le fichier Excel"
            )

            if file_path:
                # Créer un writer Excel avec un style amélioré
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Véhicules', index=False)
                    
                    # Récupérer la feuille de calcul
                    worksheet = writer.sheets['Véhicules']
                    
                    # Ajuster la largeur des colonnes
                    for idx, col in enumerate(df.columns):
                        max_length = max(
                            df[col].astype(str).apply(len).max(),
                            len(col)
                        )
                        worksheet.column_dimensions[chr(65 + idx)].width = max_length + 2

                    # Styliser l'en-tête
                    for cell in worksheet[1]:
                        cell.font = openpyxl.styles.Font(bold=True)
                        cell.fill = openpyxl.styles.PatternFill(
                            start_color='87CEFA',  # Bleu ciel
                            end_color='87CEFA',
                            fill_type='solid'
                        )

                messagebox.showinfo(
                    "Succès",
                    f"Les données ont été exportées avec succès vers:\n{file_path}"
                )
        except Exception as e:
            messagebox.showerror(
                "Erreur",
                f"Une erreur est survenue lors de l'exportation:\n{str(e)}"
            )

    def show_add_vehicle_form(self):
        # Créer une nouvelle fenêtre
        add_window = ctk.CTkToplevel(self)
        add_window.title("Ajouter un véhicule")
        add_window.geometry("400x850")

        # Variables pour stocker les entrées
        matricule_var = tk.StringVar()
        marque_var = tk.StringVar()
        modele_var = tk.StringVar()
        annee_var = tk.StringVar()
        etat_var = tk.StringVar(value="Disponible")
        kilometrage_var = tk.StringVar()
        dernieremaintenance=tk.StringVar()

        # Création du formulaire
        ctk.CTkLabel(add_window, text="Matricule:").pack(pady=5)
        matricule_entry = ctk.CTkEntry(add_window, textvariable=matricule_var)
        matricule_entry.pack(pady=5)

        ctk.CTkLabel(add_window, text="Marque:").pack(pady=5)
        marque_entry = ctk.CTkEntry(add_window, textvariable=marque_var)
        marque_entry.pack(pady=5)

        ctk.CTkLabel(add_window, text="Modèle:").pack(pady=5)
        modele_entry = ctk.CTkEntry(add_window, textvariable=modele_var)
        modele_entry.pack(pady=5)

        ctk.CTkLabel(add_window, text="Année:").pack(pady=5)
        annee_entry = ctk.CTkEntry(add_window, textvariable=annee_var)
        annee_entry.pack(pady=5)

        ctk.CTkLabel(add_window, text="État:").pack(pady=5)
        etat_combo = ttk.Combobox(add_window, textvariable=etat_var,
                                 values=["Disponible", "En mission", "En maintenance"])
        etat_combo.pack(pady=5)

        ctk.CTkLabel(add_window, text="Kilométrage:").pack(pady=5)
        kilometrage_entry = ctk.CTkEntry(add_window, textvariable=kilometrage_var)
        kilometrage_entry.pack(pady=5)
        ctk.CTkLabel(add_window, text="dernieremaintenance:").pack(pady=5)
        maint_entry = ctk.CTkEntry(add_window, textvariable=dernieremaintenance)
        maint_entry.pack(pady=5)


        # Bouton de confirmation
        def confirm_add():
            # Vérification des champs
         if not all([matricule_var.get(), marque_var.get(), modele_var.get(), 
                       annee_var.get(), etat_var.get(), kilometrage_var.get(),dernieremaintenance.get()]):
                messagebox.showwarning("Erreur", "Veuillez remplir tous les champs")
                return
         try:
            # Connexion à la base de données
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # Insertion dans la base de données
            cursor.execute("""
                INSERT INTO vehicule (id_vehicul, marque, modele, année, état, kilometrage, dernieremaintenance)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                matricule_var.get(),
                marque_var.get(),
                modele_var.get(),
                annee_var.get(),
                etat_var.get(),
                kilometrage_var.get(),
                dernieremaintenance.get(),
                datetime.now().strftime("%Y-%m-%d")
            ))
            conn.commit()
            self.vehicle_tree.insert('', 'end', values=(
                matricule_var.get(),
                marque_var.get(),
                modele_var.get(), 
                annee_var.get(),
                etat_var.get(),
                kilometrage_var.get(),
                dernieremaintenance.get(),


                datetime.now().strftime("%Y-%m-%d")
            ))
            add_window.destroy()
            messagebox.showinfo("Succès", "Véhicule ajouté avec succès")
         except sqlite3.Error as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'ajout du véhicule: {e}")
         finally:
            if 'conn' in locals():
                conn.close()
        ctk.CTkButton(add_window, text="Confirmer", command=confirm_add).pack(pady=20)

    def show_modify_vehicle_form(self):
        selected = self.vehicle_tree.selection()
        if not selected:
            messagebox.showwarning("Erreur", "Veuillez sélectionner un véhicule à modifier")
            return

        # Récupérer les données du véhicule sélectionné
        vehicle_data = self.vehicle_tree.item(selected[0])['values']

        # Créer une nouvelle fenêtre
        modify_window = ctk.CTkToplevel(self)
        modify_window.title("Modifier un véhicule")
        modify_window.geometry("400x550")

        # Variables pour stocker les entrées
        matricule_var = tk.StringVar(value=vehicle_data[0])
        marque_var = tk.StringVar(value=vehicle_data[1])
        modele_var = tk.StringVar(value=vehicle_data[2])
        annee_var = tk.StringVar(value=vehicle_data[3])
        etat_var = tk.StringVar(value=vehicle_data[4])
        kilometrage_var = tk.StringVar(value=vehicle_data[5])
        dernieremaintenance=tk.StringVar(value=vehicle_data[6])

        # Création du formulaire
        ctk.CTkLabel(modify_window, text="Matricule:").pack(pady=5)
        matricule_entry = ctk.CTkEntry(modify_window, textvariable=matricule_var)
        matricule_entry.pack(pady=5)

        ctk.CTkLabel(modify_window, text="Marque:").pack(pady=5)
        marque_entry = ctk.CTkEntry(modify_window, textvariable=marque_var)
        marque_entry.pack(pady=5)

        ctk.CTkLabel(modify_window, text="Modèle:").pack(pady=5)
        modele_entry = ctk.CTkEntry(modify_window, textvariable=modele_var)
        modele_entry.pack(pady=5)

        ctk.CTkLabel(modify_window, text="Année:").pack(pady=5)
        annee_entry = ctk.CTkEntry(modify_window, textvariable=annee_var)
        annee_entry.pack(pady=5)

        ctk.CTkLabel(modify_window, text="État:").pack(pady=5)
        etat_combo = ttk.Combobox(modify_window, textvariable=etat_var,
                                 values=["Disponible", "En mission", "En maintenance"])
        etat_combo.pack(pady=5)

        ctk.CTkLabel(modify_window, text="Kilométrage:").pack(pady=5)
        kilometrage_entry = ctk.CTkEntry(modify_window, textvariable=kilometrage_var)
        kilometrage_entry.pack(pady=5)

        ctk.CTkLabel(modify_window, text="derniere maintenance:").pack(pady=5)
        maint_entry = ctk.CTkEntry(modify_window, textvariable=dernieremaintenance)
        maint_entry.pack(pady=5)

        # Bouton de confirmation
        def confirm_modify():
            # Vérification des champs
            if not all([matricule_var.get(), marque_var.get(), modele_var.get(), 
                       annee_var.get(), etat_var.get(), kilometrage_var.get(),dernieremaintenance.get()]):
                messagebox.showwarning("Erreur", "Veuillez remplir tous les champs")
                return

            # Modifier le véhicule
            self.vehicle_tree.item(selected[0], values=(
                matricule_var.get(),
                marque_var.get(),
                modele_var.get(),
                annee_var.get(),
                etat_var.get(),
                kilometrage_var.get(),
                dernieremaintenance.get(),
                datetime.now().strftime("%Y-%m-%d")
            ))
            modify_window.destroy()
            messagebox.showinfo("Succès", "Véhicule modifié avec succès")

        ctk.CTkButton(modify_window, text="Confirmer", command=confirm_modify).pack(pady=20)

    def delete_vehicle(self):
        selected = self.vehicle_tree.selection()
        if not selected:
            messagebox.showwarning("Erreur", "Veuillez sélectionner un véhicule à supprimer")
            return

        if messagebox.askyesno("Confirmation", "Êtes-vous sûr de vouloir supprimer ce véhicule ?"):
            self.vehicle_tree.delete(selected[0])
            messagebox.showinfo("Succès", "Véhicule supprimé avec succès")



    def load_vehicle_data(self):
        """Charger les données des véhicules depuis la base existante"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
        
            cursor.execute("""
            SELECT id_vehicul, marque, modele, année, état, kilometrage, 
                   dernieremaintenance 
            FROM vehicule
             """)
            vehicles = cursor.fetchall()
            # Nettoyer les données existantes
            for item in self.vehicle_tree.get_children():
                self.vehicle_tree.delete(item)
            for vehicle in vehicles:
                self.vehicle_tree.insert('', 'end', values=vehicle)
             # Données d'exemple
            

        except sqlite3.Error as e:
             print(f"Erreur lors du chargement des véhicules: {e}")
             messagebox.showerror("Erreur", f"Erreur lors du chargement des véhicules: {e}")
        
        finally:
           if 'conn' in locals():
                 conn.close()

    def employee(self):
        self.clear_content()
        self.emp_search_var = tk.StringVar(master=self)
        self.emp_role_var = tk.StringVar(master=self, value="Tous")
        
        # Frame titre
        title_frame = ctk.CTkFrame(self.content_frame, fg_color=self.blue_ciel, height=60)
        title_frame.pack(fill="x", pady=(0, 20))
        title_label = ctk.CTkLabel(
            title_frame,
            text="Gestion des Employés",
            font=("Arial", 24, "bold"),
            text_color="white"
        )
        title_label.pack(pady=10)

        # Frame recherche
        search_frame = ctk.CTkFrame(self.content_frame)
        search_frame.pack(fill="x", padx=10, pady=5)

        # Barre de recherche
        search_label = ctk.CTkLabel(search_frame, text="Rechercher:", font=("Arial", 12))
        search_label.pack(side="left", padx=5)
        
        self.emp_search_var = tk.StringVar()
        search_entry = ctk.CTkEntry(
            search_frame,
         
            width=200,
            placeholder_text="Nom ou ID..."
        )
        search_entry.pack(side="left", padx=5)
        role_label = ctk.CTkLabel(search_frame, text="Poste:", font=("Arial", 12))
        role_label.pack(side="left", padx=5)
        
        self.emp_role_var = tk.StringVar(value="Tous")
        role_combo = ttk.Combobox(
            search_frame,
            
            values=["Tous", "Livreur", "Administrateur", "Manager", "Support client"],
            width=15,
            state="readonly"
        )
        role_combo.pack(side="left", padx=5)
        def search_employee():
           search_term = search_entry.get().lower()
           role_filter = role_combo .get()

           for item in self.emp_tree.get_children():
            values = self.emp_tree.item(item)['values']
            show = True

            if search_term:
                show = False
                for value in values[:3]:  # Recherche dans ID, Nom et Prénom
                    if str(value).lower().find(search_term) != -1:
                        show = True
                        break
            
            if role_filter != "Tous" and values[5] != role_filter:
                show = False

            if show:
                self.emp_tree.reattach(item, '', 'end')
            else:
                self.emp_tree.detach(item)

        # Filtre par poste
        

        # Boutons de recherche et réinitialisation
        search_button = ctk.CTkButton(
            search_frame,
            text="🔍 Rechercher",
            command=search_employee,
            width=100,
            fg_color=self.orange
        )
        search_button.pack(side="left", padx=5)

        reset_button = ctk.CTkButton(
            search_frame,
            text="↺ Réinitialiser",
            command=self.reset_employee_search,
            width=100,
            fg_color=self.orange
        )
        reset_button.pack(side="left", padx=5)

        # Frame pour le tableau
        table_frame = ctk.CTkFrame(self.content_frame)
        table_frame.pack(fill="both", expand=True, padx=10, pady=5)

        # Tableau des employés
        columns = ("ID", "Nom", "Prénom", "Email", "Téléphone", "Rôle", "Date d'embauche", "Statut")
        self.emp_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=15)
        
        # Configuration des colonnes
        for col in columns:
            self.emp_tree.heading(col, text=col, command=lambda c=col: self.sort_employees(c))
            width = 150 if col in ["Nom", "Prénom", "Email"] else 100
            self.emp_tree.column(col, width=width, anchor="center")

        # Scrollbar
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.emp_tree.yview)
        self.emp_tree.configure(yscrollcommand=scrollbar.set)
        
        self.emp_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Frame pour les boutons
        button_frame = ctk.CTkFrame(self.content_frame)
        button_frame.pack(fill="x", padx=10, pady=10)

        # Boutons d'action
        add_button = ctk.CTkButton(
            button_frame,
            text="➕ Ajouter",
            command=self.show_add_employee_form,
            fg_color="green",
            width=120,
            height=32
        )
        add_button.pack(side="left", padx=5)

        modify_button = ctk.CTkButton(
            button_frame,
            text="✏️ Modifier",
            command=self.show_modify_employee_form,
            fg_color="orange",
            width=120,
            height=32
        )
        modify_button.pack(side="left", padx=5)

        delete_button = ctk.CTkButton(
            button_frame,
            text="🗑️ Supprimer",
            command=self.delete_employee,
            fg_color="red",
            width=120,
            height=32
        )
        delete_button.pack(side="left", padx=5)

        export_button = ctk.CTkButton(
            button_frame,
            text="📊 Exporter Excel",
            command=self.export_to_excel_emp,
            fg_color="#2E7D32",
            width=120,
            height=32
        )
        export_button.pack(side="left", padx=5)

        # Charger les données
        self.load_employee_data()

    def show_add_employee_form(self):
        add_window = ctk.CTkToplevel(self)
        add_window.title("Ajouter un employé")
        add_window.geometry("400x600")

        # Variables
        id_var = tk.StringVar()
        nom_var = tk.StringVar()
        prenom_var = tk.StringVar()
        email_var = tk.StringVar()
        tel_var = tk.StringVar()
        poste_var = tk.StringVar(value="Livreur")
        statut_var = tk.StringVar(value="Actif")

        # Formulaire
        ctk.CTkLabel(add_window, text="ID:").pack(pady=5)
        id_entry = ctk.CTkEntry(add_window)
        id_entry.pack(pady=5)
        id_var.set(id_entry.get())

        ctk.CTkLabel(add_window, text="Nom:").pack(pady=5)
        nom_entry = ctk.CTkEntry(add_window)
        nom_entry.pack(pady=5)

        ctk.CTkLabel(add_window, text="Prénom:").pack(pady=5)
        prenom_entry = ctk.CTkEntry(add_window)
        prenom_entry.pack(pady=5)

        ctk.CTkLabel(add_window, text="Email:").pack(pady=5)
        email_entry = ctk.CTkEntry(add_window)
        email_entry.pack(pady=5)

        ctk.CTkLabel(add_window, text="Téléphone:").pack(pady=5)
        tel_entry = ctk.CTkEntry(add_window)
        tel_entry.pack(pady=5)

        ctk.CTkLabel(add_window, text="Poste:").pack(pady=5)
        poste_combo = ttk.Combobox(add_window, textvariable=poste_var,
                                  values=["Livreur", "Administrateur", "Manager", "Support client"])
        poste_combo.pack(pady=5)

        ctk.CTkLabel(add_window, text="Statut:").pack(pady=5)
        statut_combo = ttk.Combobox(add_window, textvariable=statut_var,
                                   values=["Actif", "Inactif", "En congé"])
        statut_combo.pack(pady=5)

        def confirm_add():
            if not all([id_var.get(), nom_entry.get(), prenom_entry.get(), email_entry.get(), 
                       tel_entry.get(), poste_combo.get(), statut_combo.get()]):
                messagebox.showwarning("Erreur", "Veuillez remplir tous les champs")
                return

            self.emp_tree.insert('', 'end', values=(
                id_var.get().strip(),
                nom_var.get().strip(),
                prenom_var.get().strip(),
                email_var.get().strip(),
                tel_var.get().strip(),
                poste_var.get().strip(),
                datetime.now().strftime("%Y-%m-%d"),
                statut_var.get().strip()
            ))
            add_window.destroy()
            messagebox.showinfo("Succès", "Employé ajouté avec succès")

        ctk.CTkButton(add_window, text="Confirmer", command=confirm_add).pack(pady=20)

    def show_modify_employee_form(self):
        selected = self.emp_tree.selection()
        if not selected:
            messagebox.showwarning("Erreur", "Veuillez sélectionner un employé à modifier")
            return

        emp_data = self.emp_tree.item(selected[0])['values']
        
        modify_window = ctk.CTkToplevel(self)
        modify_window.title("Modifier un employé")
        modify_window.geometry("400x600")

        # Variables
        id_var = tk.StringVar(value=emp_data[0])
        nom_var = tk.StringVar(value=emp_data[1])
        prenom_var = tk.StringVar(value=emp_data[2])
        email_var = tk.StringVar(value=emp_data[3])
        tel_var = tk.StringVar(value=emp_data[4])
        poste_var = tk.StringVar(value=emp_data[5])
        statut_var = tk.StringVar(value=emp_data[7])

        # Formulaire
        ctk.CTkLabel(modify_window, text="ID:").pack(pady=5)
        id_entry = ctk.CTkEntry(modify_window, textvariable=id_var, state="disabled")
        id_entry.pack(pady=5)

        ctk.CTkLabel(modify_window, text="Nom:").pack(pady=5)
        nom_entry = ctk.CTkEntry(modify_window, textvariable=nom_var)
        nom_entry.pack(pady=5)

        ctk.CTkLabel(modify_window, text="Prénom:").pack(pady=5)
        prenom_entry = ctk.CTkEntry(modify_window, textvariable=prenom_var)
        prenom_entry.pack(pady=5)

        ctk.CTkLabel(modify_window, text="Email:").pack(pady=5)
        email_entry = ctk.CTkEntry(modify_window, textvariable=email_var)
        email_entry.pack(pady=5)

        ctk.CTkLabel(modify_window, text="Téléphone:").pack(pady=5)
        tel_entry = ctk.CTkEntry(modify_window, textvariable=tel_var)
        tel_entry.pack(pady=5)

        ctk.CTkLabel(modify_window, text="Poste:").pack(pady=5)
        poste_combo = ttk.Combobox(modify_window, textvariable=poste_var,
                                  values=["Livreur", "Administrateur", "Manager", "Support client"])
        poste_combo.pack(pady=5)

        ctk.CTkLabel(modify_window, text="Statut:").pack(pady=5)
        statut_combo = ttk.Combobox(modify_window, textvariable=statut_var,
                                   values=["Actif", "Inactif", "En congé"])
        statut_combo.pack(pady=5)

        def confirm_modify():
            if not all([nom_var.get(), prenom_var.get(), email_var.get(), 
                       tel_var.get(), poste_var.get(), statut_var.get()]):
                messagebox.showwarning("Erreur", "Veuillez remplir tous les champs")
                return

            self.emp_tree.item(selected[0], values=(
                id_var.get(),
                nom_var.get(),
                prenom_var.get(),
                email_var.get(),
                tel_var.get(),
                poste_var.get(),
                emp_data[6],  # Garder la date d'embauche originale
                statut_var.get()
            ))
            modify_window.destroy()
            messagebox.showinfo("Succès", "Employé modifié avec succès")

        ctk.CTkButton(modify_window, text="Confirmer", command=confirm_modify).pack(pady=20)

    def delete_employee(self):
        selected = self.emp_tree.selection()
        if not selected:
            messagebox.showwarning("Erreur", "Veuillez sélectionner un employé à supprimer")
            return

        emp_data = self.emp_tree.item(selected[0])['values']
        if messagebox.askyesno("Confirmation", f"Êtes-vous sûr de vouloir supprimer l'employé {emp_data[1]} {emp_data[2]} ?"):
            self.emp_tree.delete(selected[0])
            messagebox.showinfo("Succès", "Employé supprimé avec succès")

    

    def reset_employee_search(self):
        self.emp_search_var.set("")
        self.emp_role_var.set("Tous")
        self.load_employee_dataem()

    def load_employee_data(self):
     """Charger les données des employés depuis la base de données"""
    # Nettoyer les données existantes
     for item in self.emp_tree.get_children():
        self.emp_tree.delete(item)

     try:
        # Connexion à la base de données
        conn = sqlite3.connect("suivi_coli.db")
        cursor = conn.cursor()

        # Requête SQL pour récupérer les informations des employés
        query = """
        SELECT 
            p.id_pers,
            p.nom_pers,
            p.pr_pers,
            p.email_pers,
            p.numTel_pers,
            r.nm_role,
            p.date_embch,
            p.stat_pers
        FROM personnel p
        LEFT JOIN "role de personnel" r ON p.id_role = r.id_role
        """

        cursor.execute(query)
        employees = cursor.fetchall()

        # Insérer les données dans le tableau
        for employee in employees:
            self.emp_tree.insert('', 'end', values=employee)

        # Si aucun employé trouvé
        if not employees:
            self.emp_tree.insert('', 'end', values=('Aucun employé trouvé', '', '', '', '', '', '', ''))

     except sqlite3.Error as e:
        messagebox.showerror("Erreur", f"Erreur lors du chargement des employés : {e}")
        self.emp_tree.insert('', 'end', values=('Erreur de chargement', '', '', '', '', '', '', ''))
    
     finally:
        if 'conn' in locals():
            conn.close()

    def export_to_excel_emp(self):
        try:
            data = []
            columns = ["ID", "Nom", "Prénom", "Email", "Téléphone", "Poste", "Date d'embauche", "Statut"]
            
            for item in self.emp_tree.get_children():
                values = self.emp_tree.item(item)['values']
                data.append(values)

            df = pd.DataFrame(data, columns=columns)

            file_path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Enregistrer le fichier Excel"
            )

            if file_path:
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Employés', index=False)
                    
                    worksheet = writer.sheets['Employés']
                    
                    for idx, col in enumerate(df.columns):
                        max_length = max(df[col].astype(str).apply(len).max(), len(col))
                        worksheet.column_dimensions[chr(65 + idx)].width = max_length + 2

                    for cell in worksheet[1]:
                        cell.font = openpyxl.styles.Font(bold=True)
                        cell.fill = openpyxl.styles.PatternFill(
                            start_color='87CEFA',
                            end_color='87CEFA',
                            fill_type='solid'
                        )

                messagebox.showinfo("Succès", f"Les données ont été exportées avec succès vers:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue lors de l'exportation:\n{str(e)}")

    def sort_employees(self, column):
        """Trier le tableau par colonne"""
        l = [(self.emp_tree.set(k, column), k) for k in self.emp_tree.get_children('')]
        l.sort(reverse=getattr(self, '_sort_reverse', False))
        self._sort_reverse = not getattr(self, '_sort_reverse', False)
        
        for index, (_, k) in enumerate(l):
            self.emp_tree.move(k, '', index)
    def manage_clients(self):
       
        self.clear_content()
       
        # Frame titre
        title_frame = ctk.CTkFrame(self.content_frame, fg_color=self.blue_ciel, height=60)
        title_frame.pack(fill="x", pady=(0, 20))
        title_label = ctk.CTkLabel(
            title_frame,
            text="Gestion des Clients",
            font=("Arial", 24, "bold"),
            text_color="white"
        )
        title_label.pack(pady=10)

        # Frame recherche
        search_frame = ctk.CTkFrame(self.content_frame)
        search_frame.pack(fill="x", padx=10, pady=5)

        # Barre de recherche
        search_label = ctk.CTkLabel(search_frame, text="Rechercher:", font=("Arial", 12))
        search_label.pack(side="left", padx=5)
        
       
        search_entry = ctk.CTkEntry(
            search_frame,
            width=200,
            placeholder_text="Nom ou ID..."
        )
        search_entry.pack(side="left", padx=5)
        type_label = ctk.CTkLabel(search_frame, text="Type:", font=("Arial", 12))
        type_label.pack(side="left", padx=5)
        
        
        type_combo = ttk.Combobox(
            search_frame,
            values=["Tous", "Particulier", "Professionnel", "VIP"],
            width=15,
            state="readonly"
        )
        type_combo.pack(side="left", padx=5)

        def search_client():
           search_term = search_entry.get().lower()
           type_filter = type_combo.get()

           for item in self.client_tree.get_children():
            values = self.client_tree.item(item)['values']
            show = True

            if search_term:
                show = False
                for value in values[:3]:  # Recherche dans ID, Nom et Prénom
                    if str(value).lower().find(search_term) != -1:
                        show = True
                        break
            
            if type_filter != "Tous" and values[5] != type_filter:
                show = False

            if show:
                self.client_tree.reattach(item, '', 'end')
            else:
                self.client_tree.detach(item)

        # Boutons de recherche et réinitialisation
        search_button = ctk.CTkButton(
            search_frame,
            text="🔍 Rechercher",
            command=search_client,
            width=100,
            fg_color=self.orange
        )
        search_button.pack(side="left", padx=5)
        def reset_client_search():
            search_entry.insert("0","")
            type_combo.set("Tous")
            self.load_client_data()

        reset_button = ctk.CTkButton(
            search_frame,
            text="↺ Réinitialiser",
            command=reset_client_search,
            width=100,
            fg_color=self.orange
        )
        reset_button.pack(side="left", padx=5)

        # Frame pour le tableau
        table_frame = ctk.CTkFrame(self.content_frame)
        table_frame.pack(fill="both", expand=True, padx=10, pady=5)

        # Tableau des clients
        columns = ("ID", "Nom", "Prénom", "Email", "Téléphone", "Type", "Date d'inscription", "Statut")
        self.client_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=15)
        
        # Configuration des colonnes
        for col in columns:
            self.client_tree.heading(col, text=col, command=lambda c=col: self.sort_clients(c))
            width = 150 if col in ["Nom", "Prénom", "Email"] else 100
            self.client_tree.column(col, width=width, anchor="center")

        # Scrollbar
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.client_tree.yview)
        self.client_tree.configure(yscrollcommand=scrollbar.set)
        
        self.client_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Frame pour les boutons
        button_frame = ctk.CTkFrame(self.content_frame)
        button_frame.pack(fill="x", padx=10, pady=10)

        

        

        export_button = ctk.CTkButton(
            button_frame,
            text="📊 Exporter Excel",
            command=self.export_to_excel_client,
            fg_color="#2E7D32",
            width=120,
            height=32
        )
        export_button.pack(side="left", padx=5)

        # Charger les données
        self.load_client_data()
        
    def check_tables(self):
     import sqlite3
     try:
        conn = sqlite3.connect("suivi_coli.db")
        cursor = conn.cursor()
        
        # Vérifier la table user
        cursor.execute("SELECT COUNT(*) FROM user")
        user_count = cursor.fetchone()[0]
        print(f"Nombre d'utilisateurs: {user_count}")
        
        # Vérifier la table individu
        cursor.execute("SELECT COUNT(*) FROM individu")
        individu_count = cursor.fetchone()[0]
        print(f"Nombre d'individus: {individu_count}")
        
        # Vérifier la table société
        cursor.execute("SELECT COUNT(*) FROM societé")
        societe_count = cursor.fetchone()[0]
        print(f"Nombre de sociétés: {societe_count}")
        
     except sqlite3.Error as e:
        print(f"Erreur lors de la vérification des tables: {e}")
     finally:
        if 'conn' in locals():
            conn.close()
    
    def load_client_data(self):
     import sqlite3
     # Nettoyer le tableau existant
     for item in self.client_tree.get_children():
        self.client_tree.delete(item)

     try:
        # Connexion à la base de données
        conn = sqlite3.connect("suivi_coli.db")
        cursor = conn.cursor()

        # Requête pour les individus
        cursor.execute("""
            SELECT u.id_user, 
                   i.nm_individu, 
                   i.pr_individu, 
                   i.email_individu, 
                   i.numTel_individu, 
                   'Particulier' as type_client,
                   i.ville,
                   'Actif' as statut
            FROM user u
            JOIN individu i ON u.cin_individu = i.cin_individu
            WHERE u.cin_individu IS NOT NULL
        """)
        individus = cursor.fetchall()

        # Requête pour les sociétés
        cursor.execute("""
            SELECT u.id_user,
                   s.nom_societ,
                   '' as prenom,
                   s.email_societ,
                   s.num_telsoc,
                   'Professionnel' as type_client,
                   s.ville,
                   'Actif' as statut
            FROM user u
            JOIN societé s ON u.ice_societ = s.ice_societ
            WHERE u.ice_societ IS NOT NULL
        """)
        societes = cursor.fetchall()

        # Combiner et afficher les résultats
        all_clients = individus + societes
        if all_clients:
         for client in all_clients:
            self.client_tree.insert('', 'end', values=client)

        # Si aucun client trouvé
        else:
            self.client_tree.insert('', 'end', values=('Aucun client trouvé', '', '', '', '', '', '', ''))

     except sqlite3.Error as e:
        messagebox.showerror("Erreur", f"Erreur lors du chargement des clients: {e}")
        self.client_tree.insert('', 'end', values=('Erreur de chargement', '', '', '', '', '', '', ''))
    
     finally:
        if 'conn' in locals():
            conn.close()

    def export_to_excel_client(self):
        try:
            data = []
            columns = ["ID", "Nom", "Prénom", "Email", "Téléphone", "Type", "Date d'inscription", "Statut"]
            
            for item in self.client_tree.get_children():
                values = self.client_tree.item(item)['values']
                data.append(values)

            df = pd.DataFrame(data, columns=columns)

            file_path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Enregistrer le fichier Excel"
            )

            if file_path:
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Clients', index=False)
                    
                    worksheet = writer.sheets['Clients']
                    
                    for idx, col in enumerate(df.columns):
                        max_length = max(df[col].astype(str).apply(len).max(), len(col))
                        worksheet.column_dimensions[chr(65 + idx)].width = max_length + 2

                    for cell in worksheet[1]:
                        cell.font = openpyxl.styles.Font(bold=True)
                        cell.fill = openpyxl.styles.PatternFill(
                            start_color='87CEFA',
                            end_color='87CEFA',
                            fill_type='solid'
                        )

                messagebox.showinfo("Succès", f"Les données ont été exportées avec succès vers:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue lors de l'exportation:\n{str(e)}")

    def sort_clients(self, column):
        """Trier le tableau par colonne"""
        l = [(self.client_tree.set(k, column), k) for k in self.client_tree.get_children('')]
        l.sort(reverse=getattr(self, '_sort_reverse', False))
        self._sort_reverse = not getattr(self, '_sort_reverse', False)
        
        for index, (_, k) in enumerate(l):
            self.client_tree.move(k, '', index)
    


    def tracking_packages(self):
        self.clear_content()
        
        # Frame titre
        title_frame = ctk.CTkFrame(self.content_frame, fg_color=self.blue_ciel, height=60)
        title_frame.pack(fill="x", pady=(0, 20))
        title_label = ctk.CTkLabel(
            title_frame,
            text="Suivi des Colis",
            font=("Arial", 24, "bold"),
            text_color="white"
        )
        title_label.pack(pady=10)

        # Frame de recherche
        search_frame = ctk.CTkFrame(self.content_frame)
        search_frame.pack(fill="x", padx=10, pady=5)

        # Recherche par ID
        search_label = ctk.CTkLabel(search_frame, text="Rechercher ID:", font=("Arial", 12))
        search_label.pack(side="left", padx=5)
        
        self.search_var = tk.StringVar()
        search_entry = ctk.CTkEntry(
            search_frame,
            textvariable=self.search_var,
            width=200,
            placeholder_text="Entrez l'ID..."
        )
        search_entry.pack(side="left", padx=5)
        def search_package():
            """Rechercher une demande par ID"""
            search_term = search_entry.get().upper()
            print(search_term)
        
            if not search_term:
                messagebox.showwarning("Attention", "Veuillez entrer un ID")
                return
            
        
        # Convertir l'ID recherché en minuscules pour une recherche insensible à la casse
        
            found = False
        
            for item in self.tracking_tree.get_children():
                values = self.tracking_tree.item(item)['values']
                # Convertir l'ID de la ligne en minuscules et en string pour la comparaison
                current_id = str(values[0]).upper()
                print(current_id)
            
                if current_id == search_term:
                # Sélectionner et mettre en évidence l'élément trouvé
                    self.tracking_tree.selection_set(item)
                    self.tracking_tree.see(item)
                    found = True
                    break
        
            if not found:
                pass



        # Bouton de recherche
        search_button = ctk.CTkButton(
            search_frame,
            text="🔍 Rechercher",
            command=search_package,
            width=100,
            fg_color=self.orange
        )
        search_button.pack(side="left", padx=5)

        # Frame pour le tableau
        table_frame = ctk.CTkFrame(self.content_frame)
        table_frame.pack(fill="both", expand=True, padx=10, pady=5)

        # Tableau des colis
        columns = ("ID", "Client", "Ville Départ", "Ville Arrivée", "État", "Latitude", "Longitude", "Date Livraison")
        self.tracking_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=15)
        
        # Configuration des colonnes
        for col in columns:
            self.tracking_tree.heading(col, text=col)
            self.tracking_tree.column(col, width=120, anchor="center")

        # Scrollbar
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tracking_tree.yview)
        self.tracking_tree.configure(yscrollcommand=scrollbar.set)
        
        self.tracking_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Frame pour les boutons d'action
        button_frame = ctk.CTkFrame(self.content_frame)
        button_frame.pack(fill="x", padx=10, pady=10)

        # Boutons d'action
        update_city_btn = ctk.CTkButton(
            button_frame,
            text="🌍 Modifier Ville",
            command=self.update_city_location,
            fg_color=self.orange,
            width=150
        )
        update_city_btn.pack(side="left", padx=5)

        update_status_btn = ctk.CTkButton(
            button_frame,
            text="📝 Modifier État",
            command=self.update_status,
            fg_color=self.blue_ciel,
            width=150
        )
        update_status_btn.pack(side="left", padx=5)

        # Charger les données exemple
        self.load_sample_datatr()

        
    def load_sample_datatr(self):
    # Nettoyer les données existantes
     for item in self.tracking_tree.get_children():
        self.tracking_tree.delete(item)
            
     try:
        # Connexion à la base de données
        conn = sqlite3.connect("suivi_coli.db")
        cursor = conn.cursor()

        # Requête SQL pour récupérer les données de la table colis
        query = """
        SELECT 
            id_colis,
            id_user,
            x || ', ' || y as depart,
            posit_desti_lat || ', ' || posit_desti_lon as destination,
            CASE etatlivrs
                WHEN 0 THEN 'En attente'
                WHEN 1 THEN 'En transit'
                WHEN 2 THEN 'Livré'
                ELSE 'Inconnu'
            END as etat,
            x,
            y,
            date_livrs
        FROM colis
        """

        cursor.execute(query)
        colis_data = cursor.fetchall()

        # Insérer les données dans le tableau
        for colis in colis_data:
            self.tracking_tree.insert('', 'end', values=colis)

        # Si aucune donnée trouvée
        if not colis_data:
            self.tracking_tree.insert('', 'end', values=('Aucun colis trouvé', '', '', '', '', '', '', ''))

     except sqlite3.Error as e:
        messagebox.showerror("Erreur", f"Erreur de base de données : {e}")
    
     finally:
        if 'conn' in locals():
            conn.close()

    def update_city_location(self):
        selected = self.tracking_tree.selection()
        if not selected:
            messagebox.showwarning("Attention", "Veuillez sélectionner un colis")
            return

    # Fenêtre de modification
        update_window = ctk.CTkToplevel()
        update_window.title("Modifier la ville")
        update_window.geometry("400x300")

    # Frame pour la sélection de ville
        city_frame = ctk.CTkFrame(update_window)
        city_frame.pack(fill="x", padx=10, pady=10)

        ctk.CTkLabel(city_frame, text="Sélectionner la ville:", font=("Arial", 12, "bold")).pack(pady=5)
    
        ville_var = tk.StringVar()
        ville_combo = ttk.Combobox(
        city_frame,
        textvariable=ville_var,
        values=list(VILLES_MAROC.keys()),
        state="readonly",
        width=30
    )
        ville_combo.pack(pady=5)

        def confirm_update():
            if ville_var.get():
                try:
                    lat, lng = VILLES_MAROC[ville_var.get()]
                    item = selected[0]
                    colis_id = self.tracking_tree.item(item)['values'][0]
                    conn=sqlite3.connect("suivi_coli.db")
                    c=conn.cursor()

                # Mise à jour dans la base de données
                    query = """
                    UPDATE colis 
                    SET x = ?, 
                        y = ?
                    WHERE id_colis = ?
                """
                    c.execute(query, (lat, lng, colis_id))
                    conn.commit()

                # Mise à jour dans le treeview
                    values = list(self.tracking_tree.item(item)['values'])
                    values[2] = ville_var.get()  # Ville
                    values[5] = lat  # Latitude
                    values[6] = lng  # Longitude
                    self.tracking_tree.item(item, values=values)
                
                
                    update_window.destroy()
                    conn.close()

                except sqlite3.Error as e:
                    pass

        ctk.CTkButton(
        update_window,
        text="Confirmer",
        command=confirm_update,
        fg_color="green"
    ).pack(pady=10)

    def update_status(self):
        selected = self.tracking_tree.selection()
        if not selected:
            messagebox.showwarning("Attention", "Veuillez sélectionner un colis")
            return

    # Fenêtre de modification
        status_window = ctk.CTkToplevel()
        status_window.title("Modifier l'état et la date de livraison")
        status_window.geometry("400x400")

    # Frame pour l'état
        status_frame = ctk.CTkFrame(status_window)
        status_frame.pack(fill="x", padx=10, pady=10)

        ctk.CTkLabel(status_frame, text="État:", font=("Arial", 12, "bold")).pack(pady=5)
    
        statuts = {
        "En attente": 0,
        "En transit": 1,
        "Livré": 2,
        "Retourné": 3
    }
    
        status_var = tk.StringVar()
        status_combo = ttk.Combobox(
        status_frame,
        textvariable=status_var,
        values=list(statuts.keys()),
        state="readonly",
        width=30
    )
        status_combo.pack(pady=5)

    # Frame pour la date
        date_frame = ctk.CTkFrame(status_window)
        date_frame.pack(fill="x", padx=10, pady=10)

        ctk.CTkLabel(date_frame, text="Date de livraison:", font=("Arial", 12, "bold")).pack(pady=5)
    
        date_entry = DateEntry(date_frame, width=12, background='darkblue', foreground='white', borderwidth=2)
        date_entry.pack(pady=5)

        def confirm_status():
            if status_var.get():
                try:
                    item = selected[0]
                    colis_id = self.tracking_tree.item(item)['values'][0]
                    new_status = statuts[status_var.get()]
                    new_date = date_entry.get_date().strftime("%Y-%m-%d")
                    conn=sqlite3.connect("suivi_coli.db")
                    c=conn.cursor()

                # Mise à jour dans la base de données
                    query = """
                    UPDATE colis 
                    SET etatlivrs = ?,
                        date_livrs = ?
                    WHERE id_colis = ?
                """
                    c.execute(query, (new_status, new_date, colis_id))
                    conn.commit()

                # Mise à jour dans le treeview
                    values = list(self.tracking_tree.item(item)['values'])
                    values[4] = status_var.get()  # État
                    values[7] = new_date  # Date de livraison
                    self.tracking_tree.item(item, values=values)
                
                
                    status_window.destroy()
                    conn.close()

                except sqlite3.Error as e:
                    pass

        ctk.CTkButton(
        status_window,
        text="Confirmer",
        command=confirm_status,
        fg_color="green"
    ).pack(pady=10)
    
            
    

    def sort_column(self, col):
        l = [(self.tracking_tree.set(k, col), k) for k in self.tracking_tree.get_children('')]
        l.sort(reverse=getattr(self, f'_sort_reverse_{col}', False))
        setattr(self, f'_sort_reverse_{col}', not getattr(self, f'_sort_reverse_{col}', False))
        
        for index, (_, k) in enumerate(l):
            self.tracking_tree.move(k, '', index)

    def load_sample_data(self):
        # Nettoyer les données existantes
        for item in self.tracking_tree.get_children():
            self.tracking_tree.delete(item)
            
        # Données exemple
        data = [
            ("COL001", "Ahmed Alami", "Casablanca", "Rabat", "En transit", 33.5731, -7.5898, "2024-03-15 10:00"),
            ("COL002", "Sara Bennani", "Marrakech", "Tanger", "En livraison", 31.6295, -7.9811, "2024-03-15 11:30"),
            ("COL003", "Karim Idrissi", "Fès", "Agadir", "En attente", 34.0333, -5.0000, "2024-03-15 09:15")
        ]

        # Insérer les nouvelles données
        for row in data:
            self.tracking_tree.insert('', 'end', values=row)
    def view_statistics(self):
        self.clear_content()
    
    # Frame titre
        title_frame = ctk.CTkFrame(self.content_frame, fg_color=self.blue_ciel, height=70)
        title_frame.pack(fill="x", pady=(0, 20))
    
        title_label = ctk.CTkLabel(
        title_frame,
        text="📊 Tableau de Bord",
        font=("Helvetica", 28, "bold"),
        text_color="white"
    )
        title_label.pack(pady=15)

    # Container principal
        main_container = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        main_container.pack(fill="both", expand=True, padx=20, pady=10)

        try:
            conn = sqlite3.connect("suivi_coli.db")
            cursor = conn.cursor()

        # Frame pour les KPIs
            kpi_frame = ctk.CTkFrame(main_container, fg_color="transparent")
            kpi_frame.pack(fill="x", pady=10)

        # Total des colis
            cursor.execute("SELECT COUNT(*) FROM colis")
            total_colis = cursor.fetchone()[0]

        # Nombre de retours
            cursor.execute("SELECT COUNT(*) FROM colis WHERE etatlivrs = 3")
            total_retours = cursor.fetchone()[0]

            kpi_data = [
            {
                "title": "📦 Total Colis",
                "value": str(total_colis),
                "color": self.blue_ciel
            },
            {
                "title": "❌ Retours",
                "value": str(total_retours),
                "color": self.red
            }
        ]

            for kpi in kpi_data:
                card = ctk.CTkFrame(kpi_frame, fg_color="white", corner_radius=10)
                card.pack(side="left", padx=10, fill="both", expand=True)

                title = ctk.CTkLabel(
                card,
                text=kpi["title"],
                font=("Helvetica", 14, "bold"),
                text_color=kpi["color"]
            )
                title.pack(pady=(15,5))

                value = ctk.CTkLabel(
                card,
                text=kpi["value"],
                font=("Helvetica", 24, "bold"),
                text_color="#333"
            )
                value.pack(pady=(0,15))

        # Distribution des états
            chart_frame = ctk.CTkFrame(main_container, fg_color="white", corner_radius=10)
            chart_frame.pack(fill="both", expand=True, pady=10)

            cursor.execute("""
            SELECT 
                CASE etatlivrs
                    WHEN 0 THEN 'En attente'
                    WHEN 1 THEN 'En transit'
                    WHEN 2 THEN 'Livré'
                    WHEN 3 THEN 'Retourné'
                END as etat,
                COUNT(*) as total
            FROM colis
            GROUP BY etatlivrs
        """)
            data = cursor.fetchall()

            status_data = dict(data)
        
            fig, ax = plt.subplots(figsize=(10, 6))
            colors = [self.orange, self.blue_ciel, self.green, self.red]
        
            wedges, texts, autotexts = ax.pie(
            status_data.values(),
            labels=status_data.keys(),
            colors=colors,
            autopct='%1.1f%%',
            startangle=90
        )
        
            plt.setp(autotexts, size=10, weight="bold")
            plt.setp(texts, size=10)
        
            canvas = FigureCanvasTkAgg(fig, chart_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(pady=20, padx=20, fill="both", expand=True)

        # Légende
            legend_frame = ctk.CTkFrame(main_container, fg_color="white", corner_radius=10)
            legend_frame.pack(fill="x", pady=10)
            ctk.CTkLabel(
            legend_frame,
            text="🔄 États des colis : Orange = retour | Bleu = En transit | Vert = Livré | Rouge = Retourné",
            font=("Helvetica", 12),
            text_color="#666"
        ).pack(pady=10)

        except sqlite3.Error as e:
            pass
        finally:
            if 'conn' in locals():
                conn.close()

    def logout(self):
        # Déconnexion (simulé)
        messagebox.showinfo("Déconnexion", "Vous êtes déconnecté.")

    

    def send_notifications(self):
        self.clear_content()
    
        main_container = ctk.CTkFrame(self.content_frame, fg_color="white")
        main_container.pack(fill="both", expand=True)

    # Frame titre
        title_frame = ctk.CTkFrame(main_container, fg_color=self.blue_ciel, height=70)
        title_frame.pack(fill="x", pady=(0, 20))
    
        title_label = ctk.CTkLabel(
        title_frame,
        text="🔔 Gestion des Notifications",
        font=("Helvetica", 28, "bold"),
        text_color="white"
    )
        title_label.pack(pady=15)

    # Frame pour la table
        table_frame = ctk.CTkFrame(main_container, fg_color="white")
        table_frame.pack(fill="both", expand=True, padx=20, pady=10)

    # En-tête de la table
        columns = ("ID Notification", "ID Colis", "Contenu", "Date", "Date estimée", "Horaire")
        self.notif_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=15)
    
    # Configuration des colonnes
        for col in columns:
            self.notif_tree.heading(col, text=col)
            width = 200 if col in ["Contenu"] else 100
            self.notif_tree.column(col, width=width)

        try:
            conn = sqlite3.connect("suivi_coli.db")
            cursor = conn.cursor()
        
        # Récupération des notifications
            cursor.execute("""
            SELECT 
                n.id_notif,
                n.id_colis,
                n.contenu_notif,
                n.date_notif,
                n.date_estime,
                n.horraire1 || ' - ' || n.horraire2 as horaire
            FROM notification n
            ORDER BY n.date_notif DESC
        """)
        
            notifications = cursor.fetchall()
        
            for notif in notifications:
                self.notif_tree.insert('', 'end', values=notif)

        except sqlite3.Error as e:
            pass
        finally:
            if 'conn' in locals():
                conn.close()

    # Style pour la table
        style = ttk.Style()
        style.configure("Treeview", rowheight=30, font=('Helvetica', 10))
        style.configure("Treeview.Heading", font=('Helvetica', 10, 'bold'))

        self.notif_tree.pack(fill="both", expand=True)
        self.notif_tree.bind('<Double-Button-1>', self.afficher_details_notification)

    # Scrollbar
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.notif_tree.yview)
        scrollbar.pack(side="right", fill="y")
        self.notif_tree.configure(yscrollcommand=scrollbar.set)

    # Frame pour les boutons d'action
        button_frame = ctk.CTkFrame(main_container, fg_color="white")
        button_frame.pack(fill="x", padx=20, pady=10)

    # Boutons
        ctk.CTkButton(
        button_frame,
        text="✉ Nouvelle notification",
        command=self.nouvelle_notification,
        fg_color=self.green,
        width=200
    ).pack(side="left", padx=5)

        
    def nouvelle_notification(self):
        dialog = ctk.CTkToplevel(self)
        dialog.title("Nouvelle Notification")
        dialog.geometry("600x700")
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(fg_color="white")

        main_frame = ctk.CTkFrame(dialog, fg_color="white")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)

    # Titre
        title_label = ctk.CTkLabel(
        main_frame,
        text="✉ Nouvelle Notification",
        font=("Helvetica", 20, "bold"),
        text_color=self.blue_ciel
    )
        title_label.pack(pady=10)

    # Frame formulaire
        form_frame = ctk.CTkFrame(main_frame, fg_color="white")
        form_frame.pack(fill="both", expand=True, pady=10)

    # Champs nécessaires selon la structure de la base de données
        fields = {}

    # Création des champs ID Colis et ID User
        for label in ["ID Colis", "ID User"]:
            field_frame = ctk.CTkFrame(form_frame, fg_color="transparent")
            field_frame.pack(fill="x", pady=5)

            ctk.CTkLabel(
            field_frame,
            text=f"{label} :",
            font=("Helvetica", 12, "bold"),
            width=120,
            anchor="e"
        ).pack(side="left", padx=5)

            entry = ctk.CTkEntry(field_frame, width=300)
            entry.pack(side="left", padx=5, fill="x", expand=True)

            fields[label] = entry  # Stocker directement les widgets Entry

    # Contenu de la notification
        content_frame = ctk.CTkFrame(form_frame, fg_color="transparent")
        content_frame.pack(fill="x", pady=5)

        ctk.CTkLabel(
        content_frame,
        text="Contenu :",
        font=("Helvetica", 12, "bold"),
        width=120,
        anchor="e"
    ).pack(side="left", padx=5)

        content_text = ctk.CTkTextbox(content_frame, height=100)
        content_text.pack(side="left", padx=5, fill="x", expand=True)

    # Date estimée
        date_frame = ctk.CTkFrame(form_frame, fg_color="transparent")
        date_frame.pack(fill="x", pady=10)

        ctk.CTkLabel(
        date_frame,
        text="Date estimée :",
        font=("Helvetica", 12, "bold"),
        width=120,
        anchor="e"
    ).pack(side="left", padx=5)

        date_estimee = DateEntry(date_frame, width=12, background=self.blue_ciel, foreground='white')
        date_estimee.pack(side="left", padx=5)

    # Horaires
        horaire_frame = ctk.CTkFrame(form_frame, fg_color="transparent")
        horaire_frame.pack(fill="x", pady=5)

        ctk.CTkLabel(
        horaire_frame,
        text="Horaires :",
        font=("Helvetica", 12, "bold"),
        width=120,
        anchor="e"
    ).pack(side="left", padx=5)

        horaire1_entry = ctk.CTkEntry(horaire_frame, width=100, placeholder_text="HH:MM")
        horaire1_entry.pack(side="left", padx=5)

        ctk.CTkLabel(horaire_frame, text="à").pack(side="left", padx=5)

        horaire2_entry = ctk.CTkEntry(horaire_frame, width=100, placeholder_text="HH:MM")
        horaire2_entry.pack(side="left", padx=5)

    # Fonction de sauvegarde
        def sauvegarder():
          try:
            # Récupération des valeurs directement des widgets
            id_colis = fields["ID Colis"].get().strip()
            id_user = fields["ID User"].get().strip()
            contenu = content_text.get("1.0", "end-1c").strip()
            date_est = date_estimee.get_date().strftime('%Y-%m-%d')
            h1 = horaire1_entry.get().strip()
            h2 = horaire2_entry.get().strip()

            # Validation
            if not all([id_colis, id_user, contenu, h1, h2]):
                messagebox.showwarning("Erreur", "Veuillez remplir tous les champs obligatoires")
                return

            # Connexion à la base de données
            conn = sqlite3.connect("suivi_coli.db")
            cursor = conn.cursor()

            # Insertion dans la base de données
            cursor.execute("""
                INSERT INTO notification (
                    id_colis, 
                    id_user,
                    contenu_notif,
                    date_notif,
                    date_estime,
                    horraire1,
                    horraire2
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                id_colis,
                id_user,
                contenu,
                datetime.now().strftime('%d/%m/%Y'),
                date_est,
                h1,
                h2
            ))

            conn.commit()

            # Mise à jour de l'affichage
            self.notif_tree.insert('', 'end', values=(
                cursor.lastrowid,
                id_colis,
                contenu,
                datetime.now().strftime('%Y-%m-%d'),
                date_est,
                f"{h1} - {h2}"
            ))

            dialog.destroy()

          except sqlite3.Error as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'enregistrement : {str(e)}")
          finally:
            if 'conn' in locals():
                conn.close()

    # Boutons
        button_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        button_frame.pack(fill="x", pady=20)

        ctk.CTkButton(
        button_frame,
        text="Annuler",
        command=dialog.destroy,
        fg_color=self.red,
        width=100
    ).pack(side="right", padx=5)

        ctk.CTkButton(
        button_frame,
        text="Sauvegarder",
        command=sauvegarder,
        fg_color=self.green,
        width=100
    ).pack(side="right", padx=5)

        dialog.focus_force()


    def afficher_details_notification(self, event=None):
        """Affiche les détails d'une notification dans le frame principal"""
        # Vérifier si une notification est sélectionnée
        selection = self.notif_tree.selection()
        if not selection:
            messagebox.showwarning("Attention", "Veuillez sélectionner une notification")
            return
            
        # Récupérer les données de la notification sélectionnée
        item = self.notif_tree.item(selection[0])
        colis_id = item['values'][0]
        notif = next((n for n in self.mock_notifications if n['colis_id'] == colis_id), None)
        
        if not notif:
            messagebox.showerror("Erreur", "Notification introuvable")
            return
            
        # Créer une fenêtre modale pour les détails
        details_window = self.create_modal_window("Détails de la notification", "600x800")
        
        # Frame principal
        main_frame = ctk.CTkFrame(details_window, fg_color="white") 
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        # Titre
        title_label = ctk.CTkLabel(
            main_frame,
            text="📋 Détails de la notification",
            font=("Helvetica", 20, "bold")
        )
        title_label.pack(pady=10)

        # Frame pour les détails
        details_frame = ctk.CTkFrame(main_frame, fg_color="white")
        details_frame.pack(fill="both", expand=True, pady=10)

        # Afficher les détails
        details = [
            ("ID Colis", notif['colis_id']),
            ("Destinataire", notif['destinataire']),
            ("Adresse", notif['adresse']),
            ("Téléphone", notif['telephone']),
            ("Montant", f"{notif['montant']} DH"),
            ("État", notif['etat']),
            ("Type", notif['type']),
            ("Message", notif['message']),
            ("Date création", notif['date_creation']),
            ("Livraison prévue", f"{notif['date_livraison_debut']} - {notif['date_livraison_fin']}"),
            ("Priorité", notif['priorite'])
        ]

        for label, value in details:
            detail_frame = ctk.CTkFrame(details_frame, fg_color="transparent")
            detail_frame.pack(fill="x", pady=5)
            
            ctk.CTkLabel(
                detail_frame,
                text=f"{label}:",
                font=("Helvetica", 12, "bold"),
                width=150,
                anchor="e"
            ).pack(side="left", padx=5)
            
            ctk.CTkLabel(
                detail_frame,
                text=str(value),
                font=("Helvetica", 12)
            ).pack(side="left", padx=5)

        # Bouton fermer
        ctk.CTkButton(
            main_frame,
            text="Fermer",
            command=lambda: self.close_modal(details_window),
            fg_color=self.red
        ).pack(pady=20)

    def actualiser_notifications(self):
        """Actualiser l'affichage des notifications"""
        # Effacer le tableau
        for item in self.notif_tree.get_children():
            self.notif_tree.delete(item)

        # Réinsérer les données filtrées
        filter_value = self.filter_var.get()
        
        for notif in self.mock_notifications:
            # Appliquer les filtres
            if filter_value == "Non lues" and notif['lu']:
                continue
            if filter_value == "Prioritaires" and notif['priorite'] != "Haute":
                continue
            if filter_value == "En attente" and notif['etat'] != "En attente":
                continue

            # Insérer dans le tableau
            self.notif_tree.insert('', 'end', values=(
                notif['colis_id'],
                notif['destinataire'],
                notif['type'],
                notif['etat'],
                notif['date_creation'],
                notif['priorite']
            ))

        # Si aucune donnée
        if not self.notif_tree.get_children():
            self.notif_tree.insert('', 'end', values=('Aucune notification', '', '', '', '', ''))

    def setup_search_binding(self):
        def on_search_change(*args):
            self.rechercher_notifications()
        
        self.search_var.trace_add('write', on_search_change)

    def supprimer_notification(self):
        """Supprimer une notification"""
        selection = self.notif_tree.selection()
        if not selection:
            messagebox.showwarning("Attention", "Veuillez sélectionner une notification")
            return

        if messagebox.askyesno("Confirmation", "Voulez-vous vraiment supprimer cette notification ?"):
            item = self.notif_tree.item(selection[0])
            colis_id = item['values'][0]
            
            # Supprimer des données d'exemple
            self.mock_notifications = [n for n in self.mock_notifications if n['colis_id'] != colis_id]
            
            self.actualiser_notifications()
           


    def create_modal_window(self, title, size="600x700"):
        dialog = ctk.CTkToplevel(self)
        dialog.title(title)
        dialog.geometry(size)
        dialog.transient(self)
        dialog.grab_set()
        dialog.protocol("WM_DELETE_WINDOW", lambda: self.close_modal(dialog))
        return dialog

    def close_modal(self, dialog):
        try:
            dialog.grab_release()
            dialog.destroy()
        except Exception as e:
            print(f"Erreur lors de la fermeture de la fenêtre modale: {e}")

    def clear_content(self):
        # Annuler les "after" en cours de manière sécurisée
        try:
            for after_id in self.after_ids:
                self.after_cancel(after_id)
                self.after_ids.clear()
        except Exception as e:
            print(f"Erreur lors de l'annulation des after: {e}")
        
        # Supprimer les widgets de manière sécurisée
        try:
            for widget in self.content_frame.winfo_children():
                widget.destroy()
        except Exception as e:
            print(f"Erreur lors de la suppression des widgets: {e}")

class DatabaseManager:
    def __init__(self, db_path):
        self.db_path = db_path
        
    def get_connection(self):
        return sqlite3.connect(self.db_path)
        
    def execute_query(self, query, params=None):
        try:
            with self.get_connection() as conn:
                cursor = conn.cursor()
                if params:
                    cursor.execute(query, params)
                else:
                    cursor.execute(query)
                conn.commit()
                return cursor.fetchall()
        except sqlite3.Error as e:
            print(f"Erreur SQL: {e}")
            raise
if __name__ == "__main__":
    root = ctk.CTk()
    root.geometry("1200x800")
    app = AdminPage(root)
    app.pack(fill="both", expand=True)
    root.mainloop()
